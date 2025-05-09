import sys
import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

def process_xml_to_excel(xml_file_path):
    """处理单个XML文件并转换为Excel"""
    try:
        # 解析XML文件
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        
        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "Text Data"
        
        # 设置表头
        ws['A1'] = 'ID'
        ws['B1'] = 'Content'
        
        # 设置表头样式
        header_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        
        # 遍历所有text节点
        row_num = 2
        null_rows = []
        
        for text_elem in root.findall(".//text"):
            text_id = text_elem.get('id')
            text_content = text_elem.text.strip() if text_elem.text else ""
            
            # 写入数据
            ws[f'A{row_num}'] = text_id
            ws[f'B{row_num}'] = text_content
            
            # 设置单元格样式
            cell_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            # 检查是否为%null%
            if text_content == '%null%':
                null_rows.append(row_num)
            # 检查是否为非%null%的英文内容
            elif any(c.isalpha() for c in text_content) and all(ord(c) < 128 for c in text_content if c.isalpha()):
                cell_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
            
            # 应用样式
            ws[f'A{row_num}'].fill = cell_fill
            ws[f'B{row_num}'].fill = cell_fill
            
            # 设置自动换行
            ws[f'A{row_num}'].alignment = Alignment(wrap_text=True)
            ws[f'B{row_num}'].alignment = Alignment(wrap_text=True)
            
            row_num += 1
        
        # 折叠%null%行
        if null_rows:
            null_rows.sort()
            groups = []
            current_group = [null_rows[0]]
            
            for row in null_rows[1:]:
                if row == current_group[-1] + 1:
                    current_group.append(row)
                else:
                    groups.append(current_group)
                    current_group = [row]
            
            groups.append(current_group)
            
            for group in groups:
                min_row = group[0]
                max_row = group[-1]
                ws.row_dimensions.group(min_row, max_row, hidden=True)
        
        # 调整列宽
        for col in ['A', 'B']:
            max_length = 0
            for cell in ws[col]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col].width = min(adjusted_width, 100)  # 限制最大宽度
        
        # 保存Excel文件（使用原文件名，仅更改扩展名）
        output_file = os.path.splitext(xml_file_path)[0] + '.xlsx'
        wb.save(output_file)
        return f"成功转换: {xml_file_path} -> {output_file}"
        
    except ET.ParseError as e:
        return f"解析XML文件 {xml_file_path} 时出错: XML格式错误 - {str(e)}"
    except Exception as e:
        return f"处理XML文件 {xml_file_path} 时出错: {str(e)}"

def serialize_xml(elem):
    """自定义XML序列化，严格按照指定格式输出"""
    lines = ['<?xml version="1.0" encoding="utf-8"?>', '<fmg>', '<compression>None</compression>', '<version>DarkSouls3</version>', '<bigendian>False</bigendian>', '<entries>']
    
    for child in elem.findall('entries/text'):
        text_id = child.get('id')
        text_content = child.text if child.text is not None else ""
        
        # 处理单行内容（无换行）
        if '\n' not in text_content:
            lines.append(f'<text id="{text_id}">{text_content}</text>')
        else:
            # 处理多行内容
            content_lines = text_content.split('\n')
            # 移除内容末尾的空行
            while content_lines and content_lines[-1].strip() == '':
                content_lines.pop()
            
            if not content_lines:
                lines.append(f'<text id="{text_id}"></text>')
                continue
                
            # 首行与开始标签同一行
            lines.append(f'<text id="{text_id}">{content_lines[0]}')
            
            # 中间行保持原样
            for line in content_lines[1:-1]:
                lines.append(line)
            
            # 最后一行与结束标签同一行
            if len(content_lines) > 1:
                lines.append(f'{content_lines[-1]}</text>')
    
    lines.extend(['</entries>', '</fmg>'])
    return lines

def process_excel_to_xml(excel_file_path):
    """处理单个Excel文件并转换回XML"""
    try:
        # 加载工作簿
        wb = load_workbook(excel_file_path, read_only=True)
        ws = wb.active
        
        # 创建XML结构
        fmg = ET.Element("fmg")
        
        # 添加子元素
        ET.SubElement(fmg, "compression").text = "None"
        ET.SubElement(fmg, "version").text = "DarkSouls3"
        ET.SubElement(fmg, "bigendian").text = "False"
        entries = ET.SubElement(fmg, "entries")
        
        # 从第二行开始遍历Excel数据（跳过表头）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # 遇到空行则停止处理
                break
                
            text_id = str(row[0])
            text_content = str(row[1]) if row[1] is not None else ""
            
            # 创建text节点并添加到entries
            text_elem = ET.SubElement(entries, "text")
            text_elem.set("id", text_id)
            text_elem.text = text_content
        
        # 生成XML文本
        xml_lines = serialize_xml(fmg)
        
        # 保存XML文件（使用原文件名，仅更改扩展名）
        base_name, _ = os.path.splitext(excel_file_path)
        output_file = base_name + '.xml'
        
        # 写入XML文件
        with open(output_file, 'w', encoding='utf-8') as f:
            for line in xml_lines:
                f.write(line + '\n')
        
        return f"成功转换: {excel_file_path} -> {output_file}"
        
    except Exception as e:
        return f"处理Excel文件 {excel_file_path} 时出错: {str(e)}"

def main():
    """主函数，处理拖放到程序上的所有文件"""
    results = []
    
    if len(sys.argv) < 2:
        results.append("请将XML或XLSX文件拖放到此程序上运行")
    else:
        # 获取所有拖放的文件
        files = sys.argv[1:]
        
        # 统计处理结果
        xml_count = 0
        xlsx_count = 0
        
        # 处理每个文件
        for file_path in files:
            if not os.path.isfile(file_path):
                results.append(f"跳过不存在的文件: {file_path}")
                continue
                
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext == '.xml':
                result = process_xml_to_excel(file_path)
                results.append(result)
                xml_count += 1
            elif ext == '.xlsx':
                result = process_excel_to_xml(file_path)
                results.append(result)
                xlsx_count += 1
            else:
                results.append(f"跳过不支持的文件类型: {file_path}")
        
        results.append(f"处理完成: {xml_count}个XML文件, {xlsx_count}个XLSX文件")
    
    # 打印结果
    for line in results:
        print(line)
    
    # 添加按Enter键关闭的功能
    if not sys.stdout.isatty():
        print("\n按Enter键关闭...")
        try:
            input()
        except:
            pass

if __name__ == "__main__":
    main()